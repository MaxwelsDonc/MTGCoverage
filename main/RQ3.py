import openpyxl as opxl
import os
import numpy as np


def write_xlsx(file_path, list_input_data):
    wb = opxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # # ws.append(
    #     ['model_name', 'size', 'rate of violation', 'MultiSecNeuCov',
    #      'NeuBoundCov', 'StrNeuActCov', 'TKNCov', 'TNKPat'])
    ws.append(list_input_data)
    wb.save(file_path)


def update_xlsx(file_path, list_input_data):
    if os.path.exists(file_path):
        wb = opxl.load_workbook(file_path)
        ws = wb['Sheet1']
        ws.append(list_input_data)
        wb.save(file_path)
    else:
        write_xlsx(file_path, list_input_data)


def calculate_cc(file_path, sheet_name):
    """
        函数是用来计算不同覆盖率准则之间的相关系数的函数
    Args:
        file_path: 数据的存储地方
    Returns:
    """
    if not os.path.exists(file_path):
        print("the file doesn't exist ")
        return None
    # load the excel
    rq1 = opxl.load_workbook(file_path)
    # load the sheet
    model_sheet = rq1[sheet_name]
    # load the coverage values
    KMNC = [x[0].value for x in model_sheet['D2':'D51']]
    NBC = [x[0].value for x in model_sheet['E2':'E51']]
    SNAC = [x[0].value for x in model_sheet['F2':'F51']]
    TKNC = [x[0].value for x in model_sheet['G2':'G51']]
    TKNP = [x[0].value for x in model_sheet['H2':'H51']]

    # 构造用于计算的矩阵，其中每个覆盖率准则按行存储
    coverage_value_matrix = np.array([KMNC, NBC, SNAC, TKNC, TKNP])
    cc = np.corrcoef(coverage_value_matrix)
    # 构造用于存储的数据
    sava_data = [[f'{sheet_name}', '-', '-', '-', '-', '-'], ['', 'KMNC', 'NBC', 'SNAC', 'TKNC', 'TKNP']]

    temp = ['KMNC']
    temp.extend(list(cc[0]))
    sava_data.append(temp)

    temp = ['NBC']
    temp.extend(list(cc[1]))
    sava_data.append(temp)

    temp = ['SNAC']
    temp.extend(list(cc[2]))
    sava_data.append(temp)

    temp = ['TKNC']
    temp.extend(list(cc[3]))
    sava_data.append(temp)

    temp = ['TKNP']
    temp.extend(list(cc[4]))
    sava_data.append(temp)
    # 进行存储
    for element in sava_data:
        update_xlsx('../results/RQ3.xlsx', element)


if __name__ == "__main__":
    calculate_cc('../results/RQ1.xlsx', 'LeNet1')
    calculate_cc('../results/RQ1.xlsx', 'LeNet4')
    calculate_cc('../results/RQ1.xlsx', 'LeNet5')
    calculate_cc('../results/RQ1.xlsx', 'VGG19')
    calculate_cc('../results/RQ1.xlsx', 'ResNet50')
